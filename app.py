
import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from datetime import timedelta
from openpyxl.styles import Font, Alignment, Border, Side

st.title("Kostenoverzicht generator")

uploaded_file = st.file_uploader("Upload weekoverzicht", type=["xls", "xlsx"])

if uploaded_file:
    df = pd.read_excel(uploaded_file)

    df = df[df["Project"].notna()]
    df["Project"] = df["Project"].str.strip()

    projecten = sorted(df["Project"].unique())
    project_keuze = st.selectbox("Kies project", projecten)

    # ✅ TARIEF INPUT NIEUW
    tarief = st.number_input("Tarief per m³ (€)", step=0.01, format="%.2f")

    # ✅ KBN input
    kbn = st.number_input("Tarief KBN (€)")

    brandstof_pct = ((kbn * 100) / 1100) - 100
    brandstof_pct = round(brandstof_pct, 2)

    st.write(f"Brandstoftoeslag (%): {brandstof_pct}%")

    if st.button("Genereer kostenoverzicht"):

        df_proj = df[df["Project"] == project_keuze].copy()

        gegeven_aannemer = df_proj.iloc[0]["Aannemer"] if "Aannemer" in df_proj.columns else ""

        bestandsnaam = uploaded_file.name
        jaar = bestandsnaam[0:4]
        week = bestandsnaam[4:6]
        week_text = f"WEEK {int(week)} - {jaar}"

        df_proj["volume"] = df_proj["Gelost vaste lading (m³)"]

        def time_to_hours(t):
            if pd.isna(t):
                return 0
            t = str(t).strip()
            if ":" in t:
                h, m = map(int, t.split(":"))
                return h + m / 60
            return 0

        df_proj["lostijd_uren"] = df_proj["Lostijd"].apply(time_to_hours)

        NORM = 500
        UURTARIEF = 750
        toeslag = brandstof_pct / 100

        df_proj["loskosten"] = df_proj["volume"] * tarief
        df_proj["brandstof_kosten"] = df_proj["lostijd_uren"] * 250 * toeslag
        df_proj["max_tijd"] = df_proj["volume"] / NORM

        
        import os

        base_path = os.path.dirname(__file__)
        template_path = os.path.join(base_path, "template.xlsx")

        wb = load_workbook(template_path)

        ws = wb.active

        
        # ✅ verwijder alle andere sheets behalve de actieve
        for sheet in wb.sheetnames:
            if sheet != ws.title:
                del wb[sheet]

        ws.auto_filter = None

        # ✅ ENIGE TOEVOEGING
        ws["B1"] = gegeven_aannemer
        ws["B2"] = project_keuze

        ws.oddHeader.center.text = f"LOSKOSTEN\n{week_text}"

        ws["L1"] = kbn
        ws["L1"].number_format = u'€ #,##0.00'

        ws["L2"] = brandstof_pct / 100
        ws["L2"].number_format = "0.00%"

        ws.column_dimensions["L"].width = 14
        ws.column_dimensions["O"].width = 14

        start_row = 5

        for excel_row in ws.iter_rows(min_row=start_row, max_row=500, min_col=1, max_col=15):
            for cell in excel_row:
                cell.value = None

        row_num = start_row
        totale_kosten_lijst = []

        for _, r in df_proj.iterrows():

            max_minuten = round(r["max_tijd"] * 60)
            max_tijd_td = timedelta(minutes=max_minuten)

            werkelijke_minuten = round(r["lostijd_uren"] * 60)
            extra_minuten = max(0, werkelijke_minuten - max_minuten)
            extra_tijd_td = timedelta(minutes=extra_minuten)

            extra_kosten = (extra_minuten / 60) * UURTARIEF
            totale_kosten = r["loskosten"] + extra_kosten

            totale_kosten_lijst.append(totale_kosten)

            produceren = r["volume"] / (r["lostijd_uren"]) if r["lostijd_uren"] > 0 else 0

            ws[f"A{row_num}"] = r["Datum"]
            ws[f"B{row_num}"] = r["Begeleidingsnummer"]
            ws[f"C{row_num}"] = r["Scheepsnaam"]
            ws[f"D{row_num}"] = project_keuze
            ws[f"E{row_num}"] = r["volume"]
            ws[f"F{row_num}"] = produceren
            ws[f"G{row_num}"] = tarief
            ws[f"H{row_num}"] = r["loskosten"]
            ws[f"I{row_num}"] = r["Lostijd"]
            ws[f"J{row_num}"] = toeslag
            ws[f"K{row_num}"] = r["brandstof_kosten"]

            ws[f"L{row_num}"] = max_tijd_td
            ws[f"M{row_num}"] = extra_tijd_td
            ws[f"N{row_num}"] = extra_kosten
            ws[f"O{row_num}"] = totale_kosten

            ws[f"L{row_num}"].number_format = "[h]:mm"
            ws[f"M{row_num}"].number_format = "[h]:mm"

            row_num += 1

        laatste_data_rij = row_num - 1

        if laatste_data_rij < ws.max_row:
            ws.delete_rows(laatste_data_rij + 1, ws.max_row - laatste_data_rij)

        totaal_losk = sum(totale_kosten_lijst)
        totaal_brandstof = df_proj["brandstof_kosten"].sum()
        totaal = totaal_losk + totaal_brandstof

        base_row = row_num + 1

        font_rood_bold = Font(color="FF0000", bold=True)
        align_right = Alignment(horizontal="right")

        vertical = Side(style='thin')
        horizontal = Side(style='thin')

        labels = [
            ("TOTAAL LOSKOSTEN", totaal_losk),
            ("Meerkosten brandstoftoeslag", totaal_brandstof),
            ("Te voldoen", totaal),
        ]

        for i, (tekst, waarde) in enumerate(labels):
            r = base_row + i

            cell_label = ws.cell(row=r, column=14)
            cell_label.value = tekst
            cell_label.font = font_rood_bold
            cell_label.alignment = align_right

            cell_val = ws.cell(row=r, column=15)
            cell_val.value = waarde
            cell_val.font = font_rood_bold
            cell_val.alignment = align_right
            cell_val.number_format = u'€ #,##0.00'

            cell_label.border = Border()
            cell_val.border = Border()

            # ✅ NIETS AANGEPAST
            cell_label.border = Border(right=vertical)

            if i == 2:
                for col in range(12, 16):
                    ws.cell(row=r, column=col).border = Border(
                        top=horizontal,
                        right=Side(style='thin') if col == 14 else None
                    )

        output_file = "output.xlsx"
        wb.save(output_file)

        with open(output_file, "rb") as f:
            st.download_button(
                "Download bestand",
                f,
                file_name=f"kosten_{project_keuze}.xlsx"
            )
